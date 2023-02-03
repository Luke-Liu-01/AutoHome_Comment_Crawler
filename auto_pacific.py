import time
import re
import os
import sys
from bs4 import BeautifulSoup
from selenium import webdriver
from openpyxl import Workbook
from openpyxl import load_workbook


def SetHeader(ws):
    ws.cell(row=1, column=1).value = '具体评价'
    ws.cell(row=1, column=2).value = '外观'
    ws.cell(row=1, column=3).value = '内饰'
    ws.cell(row=1, column=4).value = '空间'
    ws.cell(row=1, column=5).value = '配置'
    ws.cell(row=1, column=6).value = '动力'
    ws.cell(row=1, column=7).value = '操控'
    ws.cell(row=1, column=8).value = '油耗'
    ws.cell(row=1, column=9).value = '舒适'


def GetMouthcons(brower, ws, url):

    contents = {'外观': [], '内饰': [], '空间': [], '配置': [], '动力': [], '操控': [], '油耗': [], '舒适': []}
    brower.get(url)
    while True:
        page = brower.page_source
        bs = BeautifulSoup(page, 'html.parser')
        mouthcons = bs.find('div', class_='scollbody').find_all('div', class_='litDy clearfix')
        for mouthcon in mouthcons:
            content = mouthcon.find('div', class_='dianPing clearfix').find_all('div', class_='conLit')
            for category in content:
                tag = category.find('b').get_text().split('：')[0]
                text = category.find('span').get_text()
                try:
                    contents[tag].append(text)
                except BaseException:
                    pass
        try:
            next_page = brower.find_element_by_class_name('next')
            next_page.click()
            time.sleep(1)
        except BaseException:
            break
    SaveToExcel(ws, contents)


def SaveToExcel(ws, contents):
    category_indexs = {'外观': 2, '内饰': 3, '空间': 4, '配置': 5, '动力': 6, '操控': 7, '油耗': 8, '舒适': 9}
    row = 2

    for tag, texts in contents.items():
        for text in texts:
            try:
                ws.cell(row=row, column=1).value = text
                ws.cell(row=row, column=category_indexs[tag]).value = 1
                row += 1
            except BaseException:
                pass


if __name__ == "__main__":
    brower = webdriver.Chrome()
    wb = Workbook()
    del wb['Sheet']
    car_series = {'宝马3系': 'https://price.pcauto.com.cn/comment/sg424/', '宝马5系': 'https://price.pcauto.com.cn/comment/sg441/',
                  '宝马x1': 'https://price.pcauto.com.cn/comment/sg7209/', '宝马x3': 'https://price.pcauto.com.cn/comment/sg11244/'}
    for series_name, url in car_series.items():
        ws = wb.create_sheet(series_name)
        SetHeader(ws)
        GetMouthcons(brower, ws, url)
        wb.save('data/pacific_mouthcons.xlsx')
    brower.quit()
